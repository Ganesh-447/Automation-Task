# # import openpyxl
# # from openpyxl import load_workbook
# # from selenium import webdriver
# #
# #
# # def read_from_excel():
# #     workbook = load_workbook('testdata_ddt.xlsx')
# #     sheet = workbook.active
# #     data =[]
# #     for row in sheet.iter_rows(min_row=2,values_only=True):
# #         data.append(row)
# #     return data
import pytest
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

def get_data_excel():
    workbook = load_workbook('//tests/testdata1_ddt.xlsx')
    sheet = workbook.active
    data =[]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        data.append(row)
    return data
@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.get("https://app.vwo.com/#/login")
    yield driver
    driver.quit()

@pytest.mark.parametrize("username,password",get_data_excel())
def test_login(driver,username,password):
    email = driver.find_element(By.CSS_SELECTOR, "input[id ='login-username']")
    password = driver.find_element(By.XPATH, "//input[@id ='login-password']")
    login = driver.find_element(By.XPATH, "//button[@id ='js-login-btn']")
    email.send_keys(username)
    password.send_keys(password)
    login.click()


#
#

